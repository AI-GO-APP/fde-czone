#!/usr/bin/env python3
"""產生坤輿工業報價單主範本（含公式）。
為什麼：報價單單價會浮動、要算 5% 稅，手算易錯；範本把金額/小計/稅/含稅
合計全做成公式，使用者只填單價與數量即可自動算，避免人工計算錯誤。
輸出可直接上傳 Google Sheets（SUM/ROUND 相容）。
"""
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter

ITEM_FIRST, ITEM_LAST = 8, 17  # 品項列 8~17，共 10 列可填

wb = Workbook()
ws = wb.active
ws.title = "報價單"

# ---- 樣式 ----
thin = Side(style="thin", color="999999")
border = Border(left=thin, right=thin, top=thin, bottom=thin)
center = Alignment(horizontal="center", vertical="center")
right = Alignment(horizontal="right", vertical="center")
hdr_fill = PatternFill("solid", fgColor="E8EEF4")
title_font = Font(size=18, bold=True)
sub_font = Font(size=14, bold=True)
hdr_font = Font(bold=True)

def setc(ref, val, font=None, align=None, fill=None, bd=False, fmt=None):
    c = ws[ref]
    c.value = val
    if font: c.font = font
    if align: c.alignment = align
    if fill: c.fill = fill
    if bd: c.border = border
    if fmt: c.number_format = fmt
    return c

# ---- 抬頭 ----
ws.merge_cells("A1:E1"); setc("A1", "坤輿工業有限公司", title_font, center)
ws.merge_cells("A2:E2"); setc("A2", "報價單", sub_font, center)

# ---- 報價資訊 ----
setc("A3", "報價單號：")
setc("D3", "報價日期：")
setc("A4", "客戶名稱：")
setc("A5", "聯絡人：")
setc("D5", "電話：")

# ---- 品項表頭 ----
headers = ["品名 / 規格", "單位", "單價（未稅）", "數量", "金額（未稅）"]
for i, h in enumerate(headers):
    col = get_column_letter(i + 1)
    setc(f"{col}7", h, hdr_font, center, hdr_fill, bd=True)

# ---- 品項列：金額 = 單價 × 數量（空值時顯示空白）----
for r in range(ITEM_FIRST, ITEM_LAST + 1):
    for col in "ABCDE":
        ws[f"{col}{r}"].border = border
    ws[f"C{r}"].number_format = "#,##0.00"
    ws[f"D{r}"].number_format = "#,##0"
    setc(f"E{r}", f'=IF(OR($C{r}="",$D{r}=""),"",$C{r}*$D{r})',
         align=right, bd=True, fmt="#,##0")

# ---- 合計區（公式）----
setc(f"D{ITEM_LAST+1}", "未稅小計：", hdr_font, right)
setc(f"E{ITEM_LAST+1}", f"=SUM(E{ITEM_FIRST}:E{ITEM_LAST})", align=right, bd=True, fmt="#,##0")
setc(f"D{ITEM_LAST+2}", "營業稅（5%）：", hdr_font, right)
setc(f"E{ITEM_LAST+2}", f"=ROUND(E{ITEM_LAST+1}*0.05,0)", align=right, bd=True, fmt="#,##0")
setc(f"D{ITEM_LAST+3}", "含稅合計：", Font(bold=True), right)
setc(f"E{ITEM_LAST+3}", f"=E{ITEM_LAST+1}+E{ITEM_LAST+2}", Font(bold=True), right, bd=True, fmt="#,##0")

# ---- 備註 ----
note = ITEM_LAST + 5
setc(f"A{note}", "備註：", hdr_font)
setc(f"A{note+1}", "1. 本報價單有效期限 30 天。")
setc(f"A{note+2}", "2. 價格依市場波動調整，實際以簽約為準。")
setc(f"A{note+3}", "3. 交貨地點：買方指定地點（運費另計）。")

# ---- 表尾：報價人 / 自家資訊 ----
foot = note + 5
setc(f"A{foot}", "報價人：____________")
setc(f"A{foot+1}", "坤輿工業有限公司　統編：__________")
setc(f"A{foot+2}", "地址：________________　電話：____________")

# ---- 欄寬 ----
for col, w in {"A": 22, "B": 8, "C": 14, "D": 12, "E": 14}.items():
    ws.column_dimensions[col].width = w

out = "docs/坤輿報價單_主範本.xlsx"
wb.save(out)
print("已產生:", out)
